from openpyxl import Workbook, load_workbook
import itertools

info_name = 'Info'
tracking_name = 'Tracking'
save_dir = '/home/alex/tfg_jugglingTrackingSiteswap/AlejandroAlonso/results/excels/'

def book_initializer(system: str, ss:str) -> Workbook:
    book = Workbook()

    sheet_info = book.create_sheet(info_name)
    sheet_tracking = book.create_sheet(tracking_name)
    del book['Sheet']

    sheet_info['B2'] = 'Sistema:'
    sheet_info['C2'] = system
    sheet_info['B3'] = 'SS:'
    sheet_info['C3'] = ss
    sheet_info['B4'] = 'Num bolas:'
    sheet_info['C4'] = 'TODO'
    # TODO Añadir mas info?

    sheet_tracking.merge_cells('A1:A2')
    sheet_tracking['A1'] = 'Frame'

    return book

def obtain_column_name(num: str) -> str:
  letters = ""
  # Add 1 so that A is index 0
  num += 1
  # Convert number to stirng, every 26 values start over
  while num > 0:
    num -= 1
    letters = chr(num % 26 + 65) + letters
    num //= 26
 
  # Base case
  if len(letters) == 0:
    return "A"
 
  return letters

# Obtains the tuple represented by the ball's id
# openpyxl.utils.cell.get_column_letter(3) # returns "C"
def get_column_tuple_from_id(id: int) -> tuple[str,str]:
    return obtain_column_name(id*2-1), obtain_column_name(id*2)

def book_writer(book: Workbook, frame: int, id: int, coords: tuple[int, int]):
    sheetTracking = book[tracking_name]
    letter_1,letter_2 = get_column_tuple_from_id(id)


    if not sheetTracking[letter_1+'1'].value:
        sheetTracking.merge_cells(letter_1+'1:'+letter_2+'1')
        sheetTracking[letter_1+'1'] = f'Bola num {id}'
        sheetTracking[letter_1+'2'] = 'X'
        sheetTracking[letter_2+'2'] = 'Y'
    # TODO Igual hacer que si no se ha detectado alguna bola se ponga un guion en la fila o algo asi?
    sheetTracking[f'A{frame+2}'] = frame
    sheetTracking[letter_1+f'{frame+2}'] = coords[0]
    sheetTracking[letter_2+f'{frame+2}'] = coords[1]

def book_saver(book: Workbook, system: str, ss:str, sanitize: bool = False):
    if sanitize:
        book_sanitizer(book)
    book.save(f'{save_dir}tracking_{ss}_{system}.xlsx')
    print("Book "+ f'tracking_{ss}_{system}.xlsx' +" successfully saved in: "+ save_dir)

""" Quita frames sobrantes al principio """
def book_sanitizer(book: Workbook):
    count = 3
    while not book[tracking_name][f'A{count}'].value:
        count += 1
    book[tracking_name].delete_rows(3,count-3)
    print("Book sanitized, ", count-3, " rows deleted")

def load_data(path: str):
    book = load_workbook(path, read_only= True, data_only= True)
    sheet_info = book[info_name]
    sheetTracking = book[tracking_name]
    num_balls = sheet_info['C4'].value

    data = {}
    for b in range(0, num_balls):
        print(b)
        ball_x_column, ball_y_column = get_column_tuple_from_id(b+1)
        positions = []
        for i in range(3, sheetTracking.max_row+1):
            positions.append((sheetTracking[ball_x_column+f'{i}'].value,sheetTracking[ball_y_column+f'{i}'].value))
        data[b]=positions
    
    return data

def get_col_length(col, start=2):
    count = 0
    for cell in itertools.islice(col, start, None):
        if cell.value is not None:
            count += 1
        else:
            break

    return count

def get_first_cell_index(col):
    for idx,cell in enumerate(itertools.islice(col,2,None)):
        if cell.value is not None:
            break
    return idx+2

def get_last_cell_index(col):
    first_cell_index = get_first_cell_index(col)
    for idx,cell in enumerate(itertools.islice(col,first_cell_index,None)):
        if cell.value is None:
            break
    return idx+first_cell_index-1

def denoise(book: Workbook):
    sheetTracking = book[tracking_name]
    range_till_end = 5
    x_range = 100
    y_range = 250
    cols_seen = 1
    # Por cada columna en la hoja
    while True:
        print("out")
        for col in itertools.islice(sheetTracking.iter_cols(), cols_seen, None, 2):
            print(cols_seen)
            ended = True
            ended_inner = True
            print("iter")
            # Si termina antes de lo que le toca (teniendo en cuenta un rango)
            if (get_col_length(col) + range_till_end) < sheetTracking.max_row:
                # Compruebo con el resto de columnas
                for col2 in itertools.islice(sheetTracking.iter_cols(), cols_seen+2, None,2):
                    # Si empieza después y su primera celda está dentro de cierto rango
                    col_last_cell = get_last_cell_index(col)
                    col2_first_cell = get_first_cell_index(col2)
                    # Si alguna está vacía (ha sido movida por ejemplo) lanza excepcion
                    try:
                        if col2[col2_first_cell].row > col[col_last_cell].row:
                            pass
                        val1 = col[col_last_cell].value
                        val2 = col2[col2_first_cell].value
                        if abs(col[col_last_cell].value-col2[col2_first_cell].value)<=x_range:
                            pass
                        if abs((col[col_last_cell].offset(column=1)).value-(col2[col2_first_cell].offset(column=1)).value) <= y_range:
                            pass
                        if col2[col2_first_cell].row > col[col_last_cell].row and abs(col[col_last_cell].value-col2[col2_first_cell].value)<=x_range and abs((col[col_last_cell].offset(column=1)).value-(col2[col2_first_cell].offset(column=1)).value) <= y_range:
                            # Corto el contenido y la meto debajo mía
                            tmp = get_last_cell_index(col2)
                            range = col2[col2_first_cell].coordinate + ":" + col2[get_last_cell_index(col2)+1].offset(column=1).coordinate
                            cols=-(int(col2[0].column)-cols_seen)
                            sheetTracking.move_range(range, cols=-(int(col2[0].column)-cols_seen)+1)

                            # Borro la columna vacía
                            sheetTracking.delete_cols(col2[0].col_idx, 2)
                            
                            # Salgo para volver a comparar esa columna desde el principio y terminar de completarla si es necesario
                            ended = False
                            break
                    except:
                        pass
            if ended:
                cols_seen += 2
            else:
                break
        if ended:
            break



